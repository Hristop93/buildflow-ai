"""Excel export — the multi-sheet workbook from the prototype (SPEC 5.2 Експорт).

Builds the workbook from the latest computed snapshot, so the export is exactly
what the user saw on screen, and stamps it with the version + the date the acts
were in force (reproducibility, SPEC 8).
"""
from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session

from backend.services.sections import latest_snapshot

NAVY = "FF1F4E78"
AMBER = "FFC55A11"
_HEAD = Font(bold=True, color="FFFFFFFF")
_HEAD_FILL = PatternFill("solid", fgColor=NAVY)
_TITLE = Font(bold=True, size=14, color=NAVY)


def _sheet(wb, title):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    return ws


def _header_row(ws, row, headers, widths=None):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = _HEAD
        cell.fill = _HEAD_FILL
        cell.alignment = Alignment(horizontal="left")
    if widths:
        for c, w in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = w


def build_workbook(snapshot: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet

    summary = snapshot.get("summary", {})
    econ = snapshot.get("economics", {})
    version = snapshot.get("version_no", "?")

    # --- Резюме ---------------------------------------------------------------
    ws = _sheet(wb, "Резюме")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 24
    ws["A1"] = "Buildflow AI — обобщение"
    ws["A1"].font = _TITLE
    rows = [
        ("Брой процедури", summary.get("procedure_count")),
        ("Срок (критичен път), дни", summary.get("total_days")),
        ("Общо такси, лв", summary.get("total_fees")),
        ("CAPEX, лв", econ.get("capex")),
        ("IRR", econ.get("irr")),
        ("NPV, лв", econ.get("npv")),
        ("Присъда", econ.get("verdict")),
        ("Версия", version),
        ("Изчислено по актове в сила към", date.today().isoformat()),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)

    # --- Маршрут --------------------------------------------------------------
    ws = _sheet(wb, "Маршрут")
    _header_row(ws, 1, ["Процедура", "Институция", "Основание", "Старт (ден)", "Край (ден)", "Критична"],
                widths=[40, 30, 14, 12, 12, 10])
    for i, step in enumerate(snapshot.get("route", []), start=2):
        ws.cell(row=i, column=1, value=step.get("name"))
        ws.cell(row=i, column=2, value=step.get("institution"))
        ws.cell(row=i, column=3, value=step.get("act"))
        ws.cell(row=i, column=4, value=step.get("start_day"))
        ws.cell(row=i, column=5, value=step.get("end_day"))
        ws.cell(row=i, column=6, value="да" if step.get("is_critical") else "")

    # --- Такси ----------------------------------------------------------------
    ws = _sheet(wb, "Такси")
    _header_row(ws, 1, ["Такса", "Основа", "Ставка", "Сума, лв", "Акт", "Член"],
                widths=[36, 16, 12, 14, 18, 12])
    fees = snapshot.get("fees", {})
    row = 2
    for f in fees.get("items", []):
        cit = f.get("citation") or {}
        ws.cell(row=row, column=1, value=f.get("description"))
        ws.cell(row=row, column=2, value=f.get("basis"))
        ws.cell(row=row, column=3, value=f.get("rate"))
        ws.cell(row=row, column=4, value=f.get("amount"))
        ws.cell(row=row, column=5, value=cit.get("title"))
        ws.cell(row=row, column=6, value=cit.get("article"))
        row += 1
    ws.cell(row=row, column=1, value="ОБЩО").font = Font(bold=True)
    ws.cell(row=row, column=4, value=fees.get("total")).font = Font(bold=True)

    # --- График ---------------------------------------------------------------
    ws = _sheet(wb, "График")
    _header_row(ws, 1, ["Процедура", "Старт (ден)", "Край (ден)", "Дни", "Критична", "Статус"],
                widths=[40, 12, 12, 8, 10, 12])
    sched_nodes = snapshot.get("schedule", {}).get("nodes", {})
    ordered = sorted(sched_nodes.items(), key=lambda kv: kv[1].get("start", 0))
    for i, (_pid, n) in enumerate(ordered, start=2):
        ws.cell(row=i, column=1, value=n.get("name"))
        ws.cell(row=i, column=2, value=n.get("start"))
        ws.cell(row=i, column=3, value=n.get("end"))
        ws.cell(row=i, column=4, value=n.get("duration"))
        ws.cell(row=i, column=5, value="да" if n.get("critical") else "")
        ws.cell(row=i, column=6, value=n.get("status"))

    # --- Икономика ------------------------------------------------------------
    ws = _sheet(wb, "Икономика")
    _header_row(ws, 1, ["Година", "Паричен поток, лв", "Натрупано, лв"], widths=[10, 20, 20])
    for i, cf in enumerate(econ.get("cashflow", []), start=2):
        ws.cell(row=i, column=1, value=cf.get("year"))
        ws.cell(row=i, column=2, value=cf.get("fcf"))
        ws.cell(row=i, column=3, value=cf.get("cumulative"))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_project_xlsx(db: Session, project_id: int) -> bytes:
    """Raises NoComputedVersion if the project was never recalculated."""
    snapshot = latest_snapshot(db, project_id)
    return build_workbook(snapshot)
